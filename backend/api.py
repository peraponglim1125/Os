from flask import Flask, jsonify, request, Response
from init_db import init_db
from database import get_conn
from recognize import (
    get_best_face,
    pick_best_quality_frame,
    load_training_data,
    train_lbph,
    lbph_score_from_confidence,
    invalidate_recognizer_cache
)

import os
import time
import threading
import requests
import cv2
import numpy as np
import base64
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

app = Flask(__name__)

# =========================
# Tuning
# =========================
TARGET_WIDTH = int(os.environ.get("STREAM_WIDTH", "480"))          # 480/640
JPEG_QUALITY = int(os.environ.get("STREAM_JPEG_QUALITY", "60"))    # 55-75
DETECT_EVERY = int(os.environ.get("DETECT_EVERY", "6"))            # 4-8
PREVIEW_BRUTAL = os.environ.get("PREVIEW_BRUTAL", "false").lower() == "true"

# [NEW] ตัวแปรเก็บสถานะล่าสุดว่า "หน้าชัด/กรอบเขียว" หรือไม่
LATEST_FACE_OK = False

# =========================
# Phone camera URL
# =========================
def phone_cam_url():
    url = os.environ.get("PHONE_CAM_URL", "").strip()
    if not url:
        raise RuntimeError("PHONE_CAM_URL is not set")
    return url

def is_mjpeg_url(url: str) -> bool:
    u = (url or "").lower()
    return any(k in u for k in ["/video", "mjpeg", "stream"]) and not u.endswith((".jpg", ".jpeg", ".png"))

# =========================
# Shared frame buffer
# =========================
_latest_frame = None
_latest_ts = 0.0
_lock = threading.Lock()

def set_latest_frame(img_bgr):
    global _latest_frame, _latest_ts
    if img_bgr is None:
        return
    with _lock:
        _latest_frame = img_bgr
        _latest_ts = time.time()

def get_latest_frame_copy():
    with _lock:
        if _latest_frame is None:
            return None, 0.0
        return _latest_frame.copy(), _latest_ts

# =========================
# Utils
# =========================
def _resize_for_stream(frame):
    if frame is None:
        return None
    h, w = frame.shape[:2]
    if TARGET_WIDTH > 0 and w > TARGET_WIDTH:
        scale = TARGET_WIDTH / float(w)
        new_h = int(h * scale)
        frame = cv2.resize(frame, (TARGET_WIDTH, new_h), interpolation=cv2.INTER_AREA)
    return frame

def _decode_jpg_bytes(jpg_bytes: bytes):
    arr = np.frombuffer(jpg_bytes, dtype=np.uint8)
    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    return img

def bgr_to_b64(img_bgr):
    ok, buf = cv2.imencode(".jpg", img_bgr, [int(cv2.IMWRITE_JPEG_QUALITY), 85])
    if not ok:
        return ""
    return base64.b64encode(buf).decode("utf-8")

def encode_stream_jpg(img_bgr):
    ok, buf = cv2.imencode(".jpg", img_bgr, [int(cv2.IMWRITE_JPEG_QUALITY), JPEG_QUALITY])
    if not ok:
        return None
    return buf.tobytes()

# =========================
# Grabbers
# =========================
def snapshot_grabber(url: str):
    session = requests.Session()
    while True:
        try:
            r = session.get(url, timeout=1.5)
            if r.status_code != 200:
                time.sleep(0.05)
                continue
            img = _decode_jpg_bytes(r.content)
            if img is None:
                time.sleep(0.02)
                continue
            img = _resize_for_stream(img)
            set_latest_frame(img)
        except Exception:
            time.sleep(0.08)

def mjpeg_grabber(url: str):
    session = requests.Session()
    while True:
        try:
            r = session.get(url, stream=True, timeout=5)
            r.raise_for_status()

            byte_buf = b""
            for chunk in r.iter_content(chunk_size=4096):
                if not chunk:
                    continue
                byte_buf += chunk

                a = byte_buf.find(b"\xff\xd8")
                b = byte_buf.find(b"\xff\xd9")
                if a != -1 and b != -1 and b > a:
                    jpg = byte_buf[a:b+2]
                    byte_buf = byte_buf[b+2:]

                    img = _decode_jpg_bytes(jpg)
                    if img is None:
                        continue
                    img = _resize_for_stream(img)
                    set_latest_frame(img)
        except Exception:
            time.sleep(0.3)

# Flag สำหรับ restart grabber
_grabber_should_restart = False
_grabber_thread = None

def start_grabber():
    global _grabber_thread
    url = phone_cam_url()
    if is_mjpeg_url(url):
        _grabber_thread = threading.Thread(target=mjpeg_grabber, args=(url,), daemon=True)
        _grabber_thread.start()
    else:
        _grabber_thread = threading.Thread(target=snapshot_grabber, args=(url,), daemon=True)
        _grabber_thread.start()
    print(f"[Grabber] Started for {url}", flush=True)

def restart_grabber():
    """Reset camera - clear buffer to force reconnect"""
    global _latest_frame, _latest_ts
    with _lock:
        _latest_frame = None
        _latest_ts = 0.0
    print("[Grabber] Buffer cleared, waiting for reconnect...", flush=True)

try:
    start_grabber()
except Exception:
    pass

# =========================
# Streaming MJPEG to browser
# =========================
def gen_frames(mode="preview", student_id=""):
    global LATEST_FACE_OK

    brutal = (mode in ["enroll", "checkin"]) or PREVIEW_BRUTAL

    frame_id = 0
    last_rect = None
    last_ok = False
    last_quality = None

    while True:
        try:
            frame, ts = get_latest_frame_copy()
            if frame is None:
                time.sleep(0.05)  # รอนานขึ้นเมื่อไม่มี frame
                continue

            frame_id += 1

            if frame_id % max(DETECT_EVERY, 1) == 0:
                res = get_best_face(frame, brutal=brutal)
                if res is not None:
                    last_rect = res["rect"]
                    last_ok = bool(res.get("ok", False))
                    last_quality = res.get("quality", None)

                    LATEST_FACE_OK = last_ok
                else:
                    last_rect = None
                    last_ok = False
                    last_quality = None
                    LATEST_FACE_OK = False

            display = frame

            if last_rect is not None:
                x, y, w, h = last_rect
                color = (0, 255, 0) if last_ok else (0, 165, 255)
                cv2.rectangle(display, (x, y), (x + w, y + h), color, 2)

                if brutal and last_quality:
                    msg1 = f"sharp:{last_quality.get('blur', 0):.0f} bright:{last_quality.get('brightness', 0):.0f}"
                    msg2 = "OK - Capturable" if last_ok else "Hold still / Improve light"
                    cv2.putText(display, msg1, (10, 25), cv2.FONT_HERSHEY_SIMPLEX, 0.65, (255, 255, 255), 2)
                    cv2.putText(display, msg2, (10, 52), cv2.FONT_HERSHEY_SIMPLEX, 0.65, (255, 255, 255), 2)
                    if student_id:
                        cv2.putText(display, f"ID: {student_id}", (10, 80), cv2.FONT_HERSHEY_SIMPLEX, 0.65, (255, 255, 255), 2)
            else:
                if brutal:
                    cv2.putText(display, "NO FACE", (10, 35), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 3)

            jpg = encode_stream_jpg(display)
            if jpg is None:
                continue

            yield (
                b"--frame\r\n"
                b"Content-Type: image/jpeg\r\n\r\n" + jpg + b"\r\n"
            )
            time.sleep(0.02)  # ลด frame rate เพื่อลดโหลด
        except GeneratorExit:
            break
        except Exception:
            time.sleep(0.1)
            continue

@app.route("/video_feed")
def video_feed():
    mode = (request.args.get("mode", "preview") or "preview").strip().lower()
    sid = (request.args.get("student_id", "") or "").strip()
    if mode not in ["preview", "enroll", "checkin"]:
        mode = "preview"

    return Response(
        gen_frames(mode=mode, student_id=sid),
        mimetype="multipart/x-mixed-replace; boundary=frame"
    )

@app.route("/face_status")
def face_status():
    return jsonify({"ok": LATEST_FACE_OK})

# =========================
# Best frame source for enroll/checkin
# =========================
def fetch_frame_from_buffer():
    frame, ts = get_latest_frame_copy()
    if frame is None:
        raise RuntimeError("no frame from camera buffer (grabber)")
    return frame

def save_face_image(student_id: str) -> str:
    best = pick_best_quality_frame(fetch_frame_from_buffer, attempts=15, brutal=True)
    if best is None:
        raise RuntimeError("ไม่พบใบหน้า")

    if not best["ok"]:
        q = best["quality"]
        raise RuntimeError(f"ใบหน้าไม่ชัด/แสงไม่พอ (sharp={q['blur']:.0f}, bright={q['brightness']:.0f})")

    face_img = best["face_crop"]
    faces_dir = "/app/data/faces"
    os.makedirs(faces_dir, exist_ok=True)
    path = f"{faces_dir}/{student_id}_{int(time.time())}.jpg"
    cv2.imwrite(path, face_img)
    return path

# =========================
# APIs
# =========================
@app.route("/health")
def health():
    return jsonify({"status": "ok"})

@app.route("/reset_camera", methods=["POST"])
def reset_camera():
    """Restart the camera grabber to reconnect to phone camera"""
    try:
        restart_grabber()
        return jsonify({"status": "ok", "message": "Camera grabber restarted"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/enroll", methods=["POST"])
def enroll():
    body = request.get_json(force=True)
    sid = (body.get("student_id") or "").strip()
    name = (body.get("name") or "").strip()
    if not sid or not name:
        return jsonify({"error": "missing data"}), 400

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("INSERT INTO students(student_id, name) VALUES (?, ?)", (sid, name))
        conn.commit()
    except Exception as e:
        conn.close()
        return jsonify({"error": str(e)}), 400
    conn.close()
    return jsonify({"status": "enrolled", "student_id": sid, "name": name})

@app.route("/enroll_face", methods=["POST"])
def enroll_face():
    body = request.get_json(force=True)
    sid = (body.get("student_id") or "").strip()
    if not sid:
        return jsonify({"error": "student_id required"}), 400

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id FROM students WHERE student_id = ?", (sid,))
    s = cur.fetchone()
    if not s:
        conn.close()
        return jsonify({"error": "student not found"}), 404

    try:
        path = save_face_image(sid)
        cur.execute("INSERT INTO student_faces(student_pk, image_path) VALUES (?, ?)", (s["id"], path))
        conn.commit()
        invalidate_recognizer_cache()  # ล้าง cache เพื่อ train ใหม่รอบหน้า
    except Exception as e:
        conn.close()
        return jsonify({"error": f"failed: {e}"}), 400

    conn.close()
    return jsonify({"status": "face_enrolled", "image_path": path})

@app.route("/students")
def list_students():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT s.student_id, s.name, COUNT(f.id) as face_count
        FROM students s
        LEFT JOIN student_faces f ON f.student_pk = s.id
        GROUP BY s.id
        ORDER BY s.student_id ASC
    """)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify(rows)

@app.route("/students/<student_id>", methods=["DELETE"])
def delete_student(student_id: str):
    """ลบนักศึกษาพร้อมรูปภาพใบหน้าที่เชื่อมโยง"""
    conn = get_conn()
    cur = conn.cursor()
    
    # ค้นหา student_pk จาก student_id
    cur.execute("SELECT id FROM students WHERE student_id = ?", (student_id,))
    student = cur.fetchone()
    
    if not student:
        conn.close()
        return jsonify({"error": "ไม่พบนักศึกษา"}), 404
    
    student_pk = student["id"]
    
    # ดึง path ของรูปภาพทั้งหมดเพื่อลบไฟล์
    cur.execute("SELECT image_path FROM student_faces WHERE student_pk = ?", (student_pk,))
    face_rows = cur.fetchall()
    
    # ลบไฟล์รูปภาพจากระบบ
    for row in face_rows:
        image_path = row["image_path"]
        if image_path and os.path.exists(image_path):
            try:
                os.remove(image_path)
            except Exception as e:
                print(f"Warning: could not delete file {image_path}: {e}")
    
    # ลบข้อมูลการเช็คชื่อของนักศึกษา (ถ้าต้องการ)
    cur.execute("DELETE FROM attendance_logs WHERE student_pk = ?", (student_pk,))
    
    # ลบรูปภาพใบหน้าจากฐานข้อมูล
    cur.execute("DELETE FROM student_faces WHERE student_pk = ?", (student_pk,))
    
    # ลบข้อมูลนักศึกษา
    cur.execute("DELETE FROM students WHERE id = ?", (student_pk,))
    
    conn.commit()
    conn.close()
    
    invalidate_recognizer_cache()  # ล้าง cache เพื่อ train ใหม่รอบหน้า
    
    return jsonify({"status": "deleted", "student_id": student_id})

# =========================
# Sessions (คาบเรียน)
# =========================
@app.route("/sessions", methods=["GET"])
def list_sessions():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, title, start_at, created_at
        FROM sessions
        ORDER BY id DESC
        LIMIT 200
        """
    )
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify(rows)

@app.route("/sessions", methods=["POST"])
def create_session():
    body = request.get_json(force=True)
    title = (body.get("title") or "").strip()
    if not title:
        return jsonify({"error": "title required"}), 400

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("INSERT INTO sessions(title) VALUES (?)", (title,))
    conn.commit()
    sid = cur.lastrowid
    conn.close()
    return jsonify({"status": "created", "session_id": sid, "title": title})

@app.route("/sessions/<int:session_id>", methods=["GET"])
def get_session(session_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT id, title, start_at, created_at FROM sessions WHERE id = ?",
        (int(session_id),),
    )
    row = cur.fetchone()
    conn.close()
    if not row:
        return jsonify({"error": "session not found"}), 404
    return jsonify(dict(row))

@app.route("/sessions/<int:session_id>", methods=["DELETE"])
def delete_session(session_id: int):
    """ลบคาบเรียนพร้อมผลการเช็คชื่อทั้งหมดในคาบนั้น"""
    conn = get_conn()
    cur = conn.cursor()
    
    # ตรวจสอบว่าคาบมีอยู่จริง
    cur.execute("SELECT id, title FROM sessions WHERE id = ?", (session_id,))
    session = cur.fetchone()
    
    if not session:
        conn.close()
        return jsonify({"error": "ไม่พบคาบ"}), 404
    
    session_title = session["title"]
    
    # นับจำนวนการเช็คชื่อที่จะถูกลบ
    cur.execute("SELECT COUNT(*) as count FROM attendance_logs WHERE session_id = ?", (session_id,))
    attendance_count = cur.fetchone()["count"]
    
    # ลบผลการเช็คชื่อในคาบนี้
    cur.execute("DELETE FROM attendance_logs WHERE session_id = ?", (session_id,))
    
    # ลบคาบ
    cur.execute("DELETE FROM sessions WHERE id = ?", (session_id,))
    
    conn.commit()
    conn.close()
    
    return jsonify({
        "status": "deleted",
        "session_id": session_id,
        "title": session_title,
        "attendance_deleted": attendance_count
    })

@app.route("/attendance")
def list_attendance():
    session_id = (request.args.get("session_id") or "").strip()
    if not session_id:
        return jsonify({"error": "session_id required"}), 400

    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT student_id, name, checked_at
        FROM attendance_logs
        WHERE session_id = ?
        ORDER BY checked_at DESC
        LIMIT 200
        """,
        (int(session_id),),
    )
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return jsonify(rows)

@app.route("/attendance/export")
def export_attendance():
    """Export รายชื่อเช็คชื่อเป็นไฟล์ Excel"""
    session_id = (request.args.get("session_id") or "").strip()
    if not session_id:
        return jsonify({"error": "session_id required"}), 400

    conn = get_conn()
    cur = conn.cursor()
    
    # ดึงข้อมูลคาบ
    cur.execute("SELECT id, title, created_at FROM sessions WHERE id = ?", (int(session_id),))
    session = cur.fetchone()
    if not session:
        conn.close()
        return jsonify({"error": "session not found"}), 404
    
    session_title = session["title"]
    
    # ดึงข้อมูลการเช็คชื่อ
    cur.execute(
        """
        SELECT student_id, name, checked_at
        FROM attendance_logs
        WHERE session_id = ?
        ORDER BY checked_at ASC
        """,
        (int(session_id),),
    )
    rows = cur.fetchall()
    conn.close()

    # สร้าง Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "รายชื่อเช็คชื่อ"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title row
    ws.merge_cells('A1:D1')
    ws['A1'] = f"รายชื่อเช็คชื่อ - {session_title}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    # Header row
    headers = ["ลำดับ", "รหัสนักศึกษา", "ชื่อ-นามสกุล", "เวลาเช็คชื่อ"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for i, row in enumerate(rows, 1):
        ws.cell(row=i+3, column=1, value=i).border = thin_border
        ws.cell(row=i+3, column=2, value=row["student_id"]).border = thin_border
        ws.cell(row=i+3, column=3, value=row["name"]).border = thin_border
        ws.cell(row=i+3, column=4, value=row["checked_at"]).border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 22

    # Summary row
    summary_row = len(rows) + 5
    ws.cell(row=summary_row, column=1, value=f"รวมทั้งหมด: {len(rows)} คน")
    ws.cell(row=summary_row, column=1).font = Font(bold=True)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Create filename (ASCII only to avoid encoding issues)
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"attendance_session{session_id}_{timestamp}.xlsx"

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.route("/checkin", methods=["POST"])
def checkin():
    try:
        body = request.get_json(silent=True) or {}
        session_id = body.get("session_id", None)
        if session_id is None:
            return jsonify({"error": "session_id required"}), 400

        try:
            best = pick_best_quality_frame(fetch_frame_from_buffer, attempts=8, brutal=True)  # ลดจาก 18 เป็น 8 สำหรับ Pi 3
        except Exception as e:
            return jsonify({"error": str(e)}), 400

        if best is None:
            return jsonify({"status": "no_match", "message": "ไม่พบใบหน้า"}), 200

        display = best["rotated_img"].copy()
        x, y, w, h = best["rect"]
        cv2.rectangle(display, (x, y), (x + w, y + h), (0, 165, 255), 3)

        if not best["ok"]:
            q = best["quality"]
            cv2.putText(display, "FACE NOT CLEAR", (10, 35),
                        cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 0, 255), 3)
            return jsonify({
                "status": "no_match",
                "best_score": 0.0,
                "image": bgr_to_b64(display),
                "message": f"หน้าควรชัด/เพิ่มแสง (sharp={q['blur']:.0f}, bright={q['brightness']:.0f})"
            }), 200

        cv2.rectangle(display, (x, y), (x + w, y + h), (0, 255, 0), 3)
        img_b64 = bgr_to_b64(display)
        live_crop = best["face_crop"]

        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT s.id AS student_pk, s.student_id, s.name, f.image_path
            FROM students s
            JOIN student_faces f ON f.student_pk = s.id
            ORDER BY f.id DESC
        """)
        refs = cur.fetchall()

        images, labels = load_training_data(refs)
        if len(images) == 0:
            conn.close()
            return jsonify({
                "status": "no_match",
                "best_score": 0.0,
                "image": img_b64,
                "message": "ยังไม่มีข้อมูลใบหน้าที่ลงทะเบียน"
            }), 200

        recognizer = train_lbph(images, labels)
        if recognizer is None:
            conn.close()
            return jsonify({"status": "no_match", "best_score": 0.0, "image": img_b64}), 200

        pred_label, conf = recognizer.predict(live_crop)
        score = lbph_score_from_confidence(conf)

        # ปรับ threshold ให้เช็คง่าย (70 = ยืดหยุ่นมาก)
        pass_match = conf <= 70.0
        if not pass_match:
            conn.close()
            return jsonify({
                "status": "no_match",
                "best_score": score,
                "image": img_b64,
                "message": f"ไม่ตรงกับใคร (score={score:.2f}, dist={conf:.1f})"
            }), 200

        cur.execute("SELECT id, student_id, name FROM students WHERE id = ?", (int(pred_label),))
        s = cur.fetchone()
        if not s:
            conn.close()
            return jsonify({"status": "no_match", "best_score": score, "image": img_b64}), 200

        # กันเช็คซ้ำในคาบเดียวกัน (unique index ก็ช่วยอีกชั้น)
        cur.execute(
            """
            SELECT COUNT(*) AS c FROM attendance_logs
            WHERE student_pk = ? AND session_id = ?
            """,
            (s["id"], int(session_id)),
        )
        if cur.fetchone()["c"] > 0:
            conn.close()
            return jsonify({
                "status": "already_checked",
                "student_id": s["student_id"],
                "name": s["name"],
                "score": score,
                "image": img_b64
            }), 200

        cur.execute(
            "INSERT INTO attendance_logs(session_id, student_pk, student_id, name) VALUES (?, ?, ?, ?)",
            (int(session_id), s["id"], s["student_id"], s["name"])
        )
        conn.commit()
        conn.close()

        return jsonify({
            "status": "checked_in",
            "student_id": s["student_id"],
            "name": s["name"],
            "score": score,
            "image": img_b64
        }), 200

    except Exception as e:
        print(f"Error in checkin: {e}", flush=True)  # Add logging
        return jsonify({"error": f"Internal Server Error: {str(e)}"}), 500

# [NEW] แจ้งเตือนเมื่อเข้าผิด port
@app.route("/")
def index_root():
    return "Backend is running. Please access the Frontend at port 8080 (e.g., http://YOUR_IP:8080).", 200

if __name__ == "__main__":
    init_db()
    app.run(host="0.0.0.0", port=5000)
